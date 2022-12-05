import os
import xlrd as xl
import pandas as pd
from openpyxl import load_workbook

dir = r'C:/Users/Johnn/OneDrive/Desktop/code/count_rows/count_rows_incsv/'

files = [f for f in os.listdir(dir) if os.path.isfile(f) and f != 'main.py' and f!= ".gitignore"]

# files = ['finalized_S5cz2H2ey8Em6PPkun6FsM_aggregate_2022-12-01_scandit_localization_BJs_2022-12-01_localization_aggregate.xlsx']
count = 0
for file in files:

    wrkbk = load_workbook(file)
    sh = wrkbk.active
    total_rows = sh.max_row

    aisle_value = sh.cell(row=2, column=7).value

    for i in range(1, total_rows):
        if (sh.cell(row = i, column=7).value) != "":
            count+=1
print(f'\n🖩 Sup Playa, you loaded {len(files)} localization reports & the total count is: {count}.\n\n🤖 Stay Positive and keep grindin big dawg, we finna get there soon!\n')

# with open('C:/Users/Johnn/OneDrive/Desktop/code/count_rows/count_rows_incsv/reports', 'rs', encoding='utf-8') as report:
#     pd.read_excel(report)
#     print(report)